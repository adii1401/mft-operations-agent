import os
import sqlite3
import random
import openpyxl
import pdfplumber
from docx import Document
from datetime import datetime, timedelta
from langchain.tools import tool
import chromadb
from chromadb.utils import embedding_functions

DOCS_FOLDER = "./docs"

# ── Cache ─────────────────────────────────────────────────────────
_TP_MASTER_CACHE = None
_CHROMA_COLLECTION = None

TRANSFER_ERRORS = {
    "SFTP": ["550 Permission denied", "Connection timeout", "Authentication failed", "Host key verification failed"],
    "AS2":  ["MDN not received", "Certificate expired", "Invalid message structure", "Partner unreachable"],
    "FTPS": ["SSL handshake failed", "Passive mode error", "530 Login incorrect", "Connection reset"],
}

# ── Doc readers ───────────────────────────────────────────────────
def _read_xlsx(path):
    wb = openpyxl.load_workbook(path)
    result = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c else "" for c in row]
            else:
                row_dict = {headers[j]: row[j] for j in range(len(headers)) if j < len(row) and row[j] is not None}
                if row_dict:
                    result.append(row_dict)
    return result

def _read_txt(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def _read_pdf(path):
    text = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text.append(t)
    return "\n".join(text)

def _read_docx(path):
    doc = Document(path)
    return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

def _chunk_text(text, chunk_size=300, overlap=50):
    words = text.split()
    chunks = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i:i+chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks

# ── TP Master ─────────────────────────────────────────────────────
def load_tp_master():
    global _TP_MASTER_CACHE
    if _TP_MASTER_CACHE is not None:
        return _TP_MASTER_CACHE

    path = os.path.join(DOCS_FOLDER, "tp_master_list.xlsx")
    if not os.path.exists(path):
        return {}
    try:
        rows = _read_xlsx(path)
        tp_dict = {}
        for row in rows:
            tp_id        = str(row.get("TP ID") or "").strip().upper()
            name         = str(row.get("TP Name") or row.get("Company Name") or row.get("Name") or "")
            proto        = str(row.get("Protocol") or "")
            jo           = str(row.get("Job Owner") or "")
            jo_email     = str(row.get("JO Email") or "")
            status       = str(row.get("Status") or "Active")
            conn_type    = str(row.get("Connection Type") or "N/A")
            pwd_reset    = str(row.get("Password Reset Allowed") or "N/A")
            if tp_id:
                tp_dict[tp_id] = {
                    "name": name, "protocol": proto, "jo": jo,
                    "jo_email": jo_email, "status": status,
                    "connection_type": conn_type, "password_reset": pwd_reset
                }
        _TP_MASTER_CACHE = tp_dict
        return tp_dict
    except Exception as e:
        print(f"Error loading TP master: {e}")
        return {}

# ── ChromaDB Setup ────────────────────────────────────────────────
def get_chroma_collection():
    global _CHROMA_COLLECTION
    if _CHROMA_COLLECTION is not None:
        return _CHROMA_COLLECTION

    chroma_client = chromadb.PersistentClient(path="./chroma_db")
    ef = embedding_functions.DefaultEmbeddingFunction()
    collection = chroma_client.get_or_create_collection(
        name="mft_ops_docs",
        embedding_function=ef
    )

    existing = collection.get()
    if not existing["ids"]:
        print("Indexing docs into ChromaDB...")
        all_chunks, all_ids, all_meta = [], [], []
        chunk_id = 0

        if os.path.exists(DOCS_FOLDER):
            for filename in os.listdir(DOCS_FOLDER):
                path = os.path.join(DOCS_FOLDER, filename)
                try:
                    if filename.endswith(".txt"):
                        text = _read_txt(path)
                    elif filename.endswith(".docx"):
                        text = _read_docx(path)
                    elif filename.endswith(".pdf"):
                        text = _read_pdf(path)
                    else:
                        continue
                    for chunk in _chunk_text(text):
                        if chunk.strip():
                            all_chunks.append(chunk)
                            all_ids.append(f"chunk_{chunk_id}")
                            all_meta.append({"source": filename})
                            chunk_id += 1
                except Exception as e:
                    print(f"Could not read {filename}: {e}")

        if all_chunks:
            collection.add(documents=all_chunks, ids=all_ids, metadatas=all_meta)
            print(f"Indexed {len(all_chunks)} chunks into ChromaDB")

    _CHROMA_COLLECTION = collection
    return collection


# ── Tools ─────────────────────────────────────────────────────────

@tool
def get_tp_details(query: str) -> str:
    """Look up trading partner details by TP ID (e.g. TP001) or company name.
    Returns protocol, connection type, Job Owner name and email, password reset policy, and status."""
    tp_master = load_tp_master()
    if not tp_master:
        return "TP master list not found. Make sure tp_master_list.xlsx is in the docs/ folder."

    query_upper = query.upper().strip()

    def _format_tp(tp_id, tp):
        return (
            f"TP ID: {tp_id}\nCompany: {tp['name']}\nProtocol: {tp['protocol']}\n"
            f"Connection Type: {tp['connection_type']}\nJob Owner: {tp['jo']}\n"
            f"JO Email: {tp['jo_email']}\nPassword Reset: {tp['password_reset']}\n"
            f"Status: {tp['status']}"
        )

    if query_upper in tp_master:
        return _format_tp(query_upper, tp_master[query_upper])

    for tp_id, tp in tp_master.items():
        if query.lower() in tp["name"].lower():
            return _format_tp(tp_id, tp)

    available = ", ".join(list(tp_master.keys())[:8])
    return f"No trading partner found matching '{query}'. Available IDs: {available}"


@tool
def check_transfer_status(tp_id: str) -> str:
    """Check the latest file transfer status for a trading partner by TP ID.
    Returns last transfer time, status, file count, and error if applicable."""
    tp_master = load_tp_master()
    tp_id = tp_id.upper().strip()

    if tp_id not in tp_master:
        return f"TP ID '{tp_id}' not found in master list."

    tp = tp_master[tp_id]
    status = random.choice(["Success", "Success", "Success", "Failed", "Pending"])
    last_transfer = datetime.now() - timedelta(hours=random.randint(1, 6))
    protocol = tp.get("protocol", "SFTP")

    result = (f"TP: {tp['name']} ({tp_id})\nProtocol: {protocol}\n"
              f"Last Transfer: {last_transfer.strftime('%Y-%m-%d %H:%M')}\n"
              f"Status: {status}\nFiles Processed: {random.randint(1, 20)}\n")

    if status == "Failed":
        error = random.choice(TRANSFER_ERRORS.get(protocol, ["Unknown error"]))
        result += f"Error: {error}\nAction: Check MFT logs, verify connection settings, escalate to L2 if unresolved in 2 hours."
    elif status == "Pending":
        result += "Note: Transfer pending over 30 minutes. Check BIS Process Monitor."

    return result


@tool
def get_pending_followups(filter_type: str = "all") -> str:
    """Get follow-up items from the tracker database.
    filter_type options: 'all', 'pending', 'overdue', 'escalated'"""
    try:
        conn = sqlite3.connect("../mft-email-responder/data/followups.db")
        cursor = conn.cursor()
        cursor.execute("SELECT subject, sender, sent_at, priority, deadline, status FROM followups ORDER BY deadline ASC")
        rows = cursor.fetchall()
        conn.close()
    except Exception:
        rows = [
            ("SFTP failure TP001", "user@company.com", "2026-03-01 10:00", "P2 - Degraded Service", "2026-03-01 14:00", "Pending"),
            ("Password reset TP003", "user@company.com", "2026-03-01 09:00", "P3 - Non-critical", "2026-03-02 09:00", "Escalated"),
        ]

    if not rows:
        return "No follow-up items found."

    now = datetime.now()
    result = []
    for subject, sender, sent_at, priority, deadline, status in rows:
        try:
            overdue = now > datetime.strptime(deadline, "%Y-%m-%d %H:%M") and status == "Pending"
        except Exception:
            overdue = False
        if filter_type == "pending" and status != "Pending":
            continue
        if filter_type == "overdue" and not overdue:
            continue
        if filter_type == "escalated" and status != "Escalated":
            continue
        tag = " OVERDUE" if overdue else ""
        result.append(f"[{status}{tag}] {subject} | {priority} | Deadline: {deadline}")

    return "\n".join(result) if result else f"No {filter_type} follow-ups found."


@tool
def search_knowledge_base(query: str) -> str:
    """Search MFT/EDI knowledge base using semantic vector search.
    Finds relevant SOPs, procedures, and past resolved cases from MFT documentation."""
    try:
        collection = get_chroma_collection()
        results = collection.query(
            query_texts=[query],
            n_results=3,
            include=["documents", "metadatas", "distances"]
        )
        chunks = results["documents"][0]
        metadatas = results["metadatas"][0]
        distances = results["distances"][0]

        if not chunks:
            return "No relevant procedures found in knowledge base."

        output = []
        for chunk, meta, dist in zip(chunks, metadatas, distances):
            confidence = round(max(0, 1 - dist / 2) * 100, 1)
            output.append(f"[Source: {meta['source']} | Confidence: {confidence}%]\n{chunk}")

        return "\n\n---\n\n".join(output)

    except Exception as e:
        return f"Knowledge base search error: {str(e)}"


@tool
def draft_escalation_email(tp_id: str, issue_description: str) -> str:
    """Draft a professional escalation email for a trading partner issue.
    Provide TP ID and a brief description of the issue."""
    tp_master = load_tp_master()
    tp_id = tp_id.upper().strip()

    if tp_id not in tp_master:
        return f"TP ID '{tp_id}' not found. Cannot draft escalation email."

    tp = tp_master[tp_id]
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    pwd_note = (
        f"\n⚠ Password Reset Policy: {tp['password_reset']}"
        if "approval" in tp.get("password_reset", "").lower() or "escalate" in tp.get("password_reset", "").lower()
        else ""
    )

    return f"""Subject: [ESCALATION] {tp['protocol']} Transfer Issue — {tp['name']} ({tp_id})

Dear {tp['jo']},

I am writing to escalate an ongoing issue with {tp['name']} ({tp_id}) that requires your attention.

Issue: {issue_description}
Reported At: {now}

Standard troubleshooting steps have been completed without resolution.

TP Details: {tp_id} | {tp['name']} | {tp['protocol']} | {tp['connection_type']}{pwd_note}

Please advise on business priority and authorize any configuration changes required.

Best regards,
MFT Support Team
CC: {tp['jo_email']}"""


@tool
def generate_onboarding_checklist(protocol: str, tp_name: str) -> str:
    """Generate a step-by-step onboarding checklist for a new trading partner.
    Provide protocol (SFTP/AS2/FTPS) and trading partner name."""
    protocol = protocol.upper().strip()

    common = [
        f"1. Collect TP details — company name, contact email, technical contact",
        f"2. Get Job Owner (JO) sign-off before starting setup",
        f"3. Create TP profile in Seeburger BIS — name: {tp_name}",
        f"4. Configure {protocol} connection parameters",
        f"5. Set up file naming conventions and directory paths",
        f"6. Configure error notifications and alerting",
        f"7. Run test transfer and confirm successful acknowledgement",
        f"8. Get JO sign-off on successful test before go-live",
        f"9. Update TP master list and internal documentation",
        f"10. Notify TP of go-live date and provide support contact",
    ]

    protocol_specific = {
        "SFTP": ["- Exchange SSH keys or configure password credentials",
                 "- Confirm port 22 is accessible, whitelist MFT server IP",
                 "- Verify SFTP directory structure and permissions"],
        "AS2":  ["- Exchange AS2 certificates with trading partner",
                 "- Configure AS2 ID, URL, port (default 4080)",
                 "- Set up MDN settings and agree on encryption algorithms"],
        "FTPS": ["- Install partner SSL certificate in BIS trust store",
                 "- Configure passive mode IP and port range",
                 "- Confirm firewall allows control (21) and data channels"],
    }

    result = f"ONBOARDING CHECKLIST — {tp_name} ({protocol})\n{'='*50}\n\n"
    result += "\n".join(common)
    result += f"\n\n{protocol}-SPECIFIC:\n"
    result += "\n".join(protocol_specific.get(protocol, ["Follow standard connection setup procedure"]))
    result += "\n\nEstimated setup time: 3-5 business days"
    return result


@tool
def detect_sla_breaches(filter_status: str = "all") -> str:
    """Detect trading partners with SLA transfer breaches or at-risk transfers.
    filter_status options: 'all', 'breached' (past SLA threshold), 'at_risk' (within 1hr of breach)"""
    tp_master = load_tp_master()
    if not tp_master:
        return "TP master list not available."

    now = datetime.now()
    SLA_THRESHOLDS = {"SFTP": 4, "AS2": 2, "FTPS": 6}
    results = []

    for tp_id, tp in tp_master.items():
        protocol = tp.get("protocol", "SFTP")
        threshold_hours = SLA_THRESHOLDS.get(protocol, 4)
        random.seed(f"{tp_id}{now.strftime('%Y-%m-%d %H')}")
        hours_since = round(random.uniform(0.5, threshold_hours + 3), 1)
        last_transfer = now - timedelta(hours=hours_since)
        hours_remaining = threshold_hours - hours_since

        if hours_remaining < 0:
            status = "BREACHED"
            tag = f"SLA breached by {abs(hours_remaining):.1f}h"
        elif hours_remaining <= 1:
            status = "AT RISK"
            tag = f"SLA breach in {hours_remaining:.1f}h"
        else:
            status = "OK"
            tag = f"{hours_remaining:.1f}h remaining"

        if filter_status == "breached" and status != "BREACHED":
            continue
        if filter_status == "at_risk" and status != "AT RISK":
            continue

        results.append(
            f"[{status}] {tp_id} — {tp['name']} | {protocol} | "
            f"Last transfer: {last_transfer.strftime('%Y-%m-%d %H:%M')} | {tag} | JO: {tp['jo']}"
        )

    if not results:
        return f"No {filter_status} SLA issues found."

    header = f"SLA BREACH REPORT ({filter_status.upper()}) — {now.strftime('%Y-%m-%d %H:%M')}\n{'='*60}\n"
    return header + "\n".join(results)


@tool
def get_onboarding_status(query: str = "all") -> str:
    """Get onboarding status for new trading partners being set up.
    query: 'all' to list all onboarding TPs, or provide a TP ID (e.g. TP009) or partial name.
    Stages: Not Started → In Progress → Testing → Go-Live → Complete"""
    path = os.path.join(DOCS_FOLDER, "onboarding_tracker.xlsx")
    if not os.path.exists(path):
        return "Onboarding tracker not found. Make sure onboarding_tracker.xlsx is in the docs/ folder."

    try:
        records = _read_xlsx(path)
    except Exception as e:
        return f"Error reading onboarding tracker: {str(e)}"

    if not records:
        return "No onboarding records found in tracker."

    query = query.strip()
    now = datetime.now()

    if query.lower() != "all":
        q = query.upper()
        records = [
            r for r in records
            if q == str(r.get("TP ID", "")).upper() or query.lower() in str(r.get("TP Name", "")).lower()
        ]

    if not records:
        return f"No onboarding records found matching '{query}'."

    STAGE_ORDER = ["Not Started", "In Progress", "Testing", "Go-Live", "Complete"]
    lines = [f"ONBOARDING STATUS REPORT — {now.strftime('%Y-%m-%d %H:%M')}\n{'='*60}"]

    for r in records:
        try:
            raw_date = r.get("Target Go-Live", "")
            if hasattr(raw_date, "date"):
                target = raw_date.replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                target = datetime.strptime(str(raw_date).split(" ")[0], "%Y-%m-%d")
            days_left = (target - now).days
            overdue_flag = " ⚠ OVERDUE" if days_left < 0 else f" ({days_left}d to go-live)"
        except Exception:
            overdue_flag = ""

        stage = str(r.get("Stage", "Unknown"))
        stage_idx = STAGE_ORDER.index(stage) + 1 if stage in STAGE_ORDER else "?"
        progress = f"Step {stage_idx}/{len(STAGE_ORDER)}"

        lines.append(
            f"\nTP ID     : {r.get('TP ID', 'N/A')}\n"
            f"Name      : {r.get('TP Name', 'N/A')}\n"
            f"Protocol  : {r.get('Protocol', 'N/A')}\n"
            f"Stage     : {stage} ({progress}){overdue_flag}\n"
            f"Assigned  : {r.get('Assigned To', 'N/A')}\n"
            f"Started   : {str(r.get('Started', 'N/A')).split(' ')[0]}\n"
            f"Go-Live   : {str(r.get('Target Go-Live', 'N/A')).split(' ')[0]}\n"
            f"Notes     : {r.get('Notes', 'N/A')}"
        )

    return "\n".join(lines)